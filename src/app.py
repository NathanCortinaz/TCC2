# bibliotecas para interface front e backend
from __future__ import unicode_literals
from distutils.log import debug
from tempfile import SpooledTemporaryFile
import eel
import os

# bibliotecas para análise facial
from fer import FER
from time import sleep, perf_counter
import pandas as pd
import cv2

# biblioteca para acessar explorador de arquivo
import subprocess

# bibliotecas para gráficos
import collections
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import openpyxl
import numpy

# definição do caminho frontend
eel.init(f'{os.path.dirname(os.path.realpath(__file__))}/web')
path = f'{os.path.dirname(os.path.realpath(__file__))}\Results'


class FaceAnalyzer():
    def __init__(self, capture):
        self.capture = capture
        if not self.capture.isOpened():
            eel.camError()()
            exit()
        else:
            self.start = perf_counter()
            self.current_reaction = 'neutro'
            self.new_reaction = 'neutro'
            self.face_found = False
            self.reactions = [(self.current_reaction, 0)]
            self.reactions_step = []
            self.sorted_frequency_list = []
            self.detector = FER()
            self.faceCascade = cv2.CascadeClassifier(
                cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

    def detect_reaction(self):
        '''Verifica video da webcam buscando por expressões faciais'''

        success, self.frame = self.capture.read()
        reaction, score = self.detector.top_emotion(self.frame)
        if reaction != None:
            reaction = self.translate_reactions(reaction)
            self.new_reaction = reaction
            self.face_found = True
        else:
            print("Face não identificada...")
            self.face_found = False
        print(f'{self.new_reaction=}')

    def show_faces(self):
        '''Exibe faces detectadas na webcam com suas expressões'''

        if self.face_found == True:
            faces = self.faceCascade.detectMultiScale(self.frame, 1.1, 4)
            for (x, y, w, h) in faces:
                cv2.rectangle(self.frame, (x, y), (x+w, y+h), (0, 255, 0), 2)
                cv2.putText(self.frame, self.new_reaction, (x, y-10),
                            cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2, cv2.LINE_4)
            cv2.startWindowThread()
            cv2.imshow('Analisador de Expressões', self.frame)
            cv2.waitKey(1)

    def check_reaction(self):
        '''Verifica se há nova reação e adiciona à lista de reações'''

        if self.new_reaction != self.current_reaction:

            self.current_reaction = self.new_reaction
            running_time = perf_counter() - self.start
            self.reactions.append((self.new_reaction, running_time))

    def check_stop(self, debug_mode):
        '''Verifica se usuário deseja parar e salva reações no arquivo'''

        if debug_mode == True:
            exitApp = eel.checkDebugButton()()
        else:
            exitApp = eel.checkRunButton()()
        if exitApp:
            running_time = perf_counter() - self.start
            self.reactions.append((self.current_reaction, running_time))
            self.create_step_reactions_list()
            self.create_frequency_list()
            print(f'\n{self.reactions = }')
            print(f'\n{self.reactions_step = }')
            print(f'\n{self.sorted_frequency_list = }')
            self.save_to_excel()
            return True

    def create_step_reactions_list(self):
        '''Cria lista com reações por segundo'''

        current_reaction = 'neutral'
        step = 0
        for reaction in self.reactions:
            while step < reaction[1]:
                self.reactions_step.append((current_reaction, step))
                step += 1
            current_reaction = reaction[0]
        self.reactions_step.append((reaction[0], step))

    def create_frequency_list(self):
        '''Cria lista com frequência de reações'''

        reactions_list = [reaction[0] for reaction in self.reactions_step]
        frequency_list = list(collections.Counter(reactions_list).items())
        self.sorted_frequency_list = sorted(frequency_list, key=lambda x: (-x[1], x[0]))

    def translate_reactions(self, reaction):
        '''Traduz lista de reações'''
        match reaction:
            case 'neutral':
                reaction = 'neutro'
            case 'happy':
                reaction = 'feliz'
            case 'fear':
                reaction = 'medo'
            case 'angry':
                reaction = 'raiva'
            case 'sad':
                reaction = 'triste'
            case 'disgust':
                reaction = 'desgosto'
            case 'surprise':
                reaction = 'surpresa'
        return reaction

    def save_to_excel(self):
        '''Gera gráficos e salva junto com dados em Excel'''

        df1 = pd.DataFrame(self.reactions_step)
        x1 = df1[1]
        y1 = df1[0]
        df2 = pd.DataFrame(self.sorted_frequency_list)
        x2 = df2[0]
        y2 = df2[1]

        width = 5 if len(self.reactions_step)/6 < 5 else len(self.reactions_step)/6

        fig = plt.figure()
        fig = plt.figure(figsize=(width, 5))
        ax = fig.gca()
        plt.plot(x1, y1, linewidth=2, marker=".", markersize=10)
        plt.ylabel('Reação', fontweight='bold')
        plt.xlabel('Tempo [s]', fontweight='bold')
        ax.set_xticks(numpy.arange(0, self.reactions_step[-1][1]+2, 2))
        plt.grid(linestyle=':')
        plt.tight_layout()
        plt.savefig("GraficoLinhas.png")
        plt.clf()

        plt.figure()
        plt.bar(x2, y2)
        plt.xlabel('Reação', fontweight='bold')
        plt.ylabel('Tempo [s]', fontweight='bold')
        for i in range(len(x2)):
            plt.text(i, y2[i], y2[i], ha='center')
        plt.tight_layout()
        plt.savefig("GraficoBarras.png")
        plt.clf()

        wb = openpyxl.Workbook()

        sheet = wb.active
        sheet.title = "Emoções por segundo"
        sheet.cell(1, 1).value = "REAÇÃO"
        sheet.cell(1, 2).value = "TEMPO [s]"
        for item in self.reactions_step:
            sheet.append(item)
        grafico_linhas = openpyxl.drawing.image.Image("GraficoLinhas.png")
        grafico_linhas.anchor = 'D1'
        sheet.add_image(grafico_linhas)

        sheet = wb.create_sheet()
        sheet.title = 'Ranking de reações'
        sheet.cell(1, 1).value = "REAÇÃO"
        sheet.cell(1, 2).value = "TEMPO [s]"
        for item in self.sorted_frequency_list:
            sheet.append(item)
        grafico_barras = openpyxl.drawing.image.Image("GraficoBarras.png")
        grafico_barras.anchor = 'D1'
        sheet.add_image(grafico_barras)

        wb.save(f"{path}/Resultados.xlsx")
        wb.close

        os.remove("GraficoLinhas.png")
        os.remove("GraficoBarras.png")


def start(debug_mode):
    '''Inicia aplicação'''

    capture = cv2.VideoCapture(0)
    fa = FaceAnalyzer(capture)
    os.system('cls' if os.name == 'nt' else 'clear')
    while True:
        fa.detect_reaction()
        if debug_mode == True:
            fa.show_faces()
        fa.check_reaction()
        stop_running = fa.check_stop(debug_mode)
        if stop_running == True:
            break
        sleep(1)
    capture.release()
    cv2.destroyAllWindows()


@eel.expose
def run():
    start(debug_mode=False)


@eel.expose
def debug():
    start(debug_mode=True)


@eel.expose
def open_results():
    '''Abre pasta contendo arquivo de resultados'''

    subprocess.Popen(f'explorer "{path}"')


os.system('cls' if os.name == 'nt' else 'clear')
eel.start('index.html', size=(720, 235))
